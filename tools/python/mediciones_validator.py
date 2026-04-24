#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
mediciones_validator.py — Cruce programático bc3 vs Excel
==========================================================
Compara las cantidades del presupuesto bc3 con las del Excel de mediciones.
Sin LLM: todos los valores se leen determinísticamente.

Fuente de verdad para diferencias: Excel (Excel > bc3).

Uso:
  python3 mediciones_validator.py presupuesto.bc3 mediciones.xlsx \\
      --sheet=HOJA \\
      --col-code=B    (columna con el código de partida)
      --col-qty=H     (columna con la cantidad)
      [--col-unit=C]  (columna con la unidad — opcional)
      [--col-desc=D]  (columna con la descripción — opcional, para el informe)
      [--header=2]    (fila de cabecera, defecto 1 — las de datos empiezan en header+1)
      [--tolerance=0.01]
      [--output=informe.txt]

Ejemplos:
  python3 tools/mediciones_validator.py \\
      presupuesto.bc3 "535.2.2 Control-Calidad.xlsx" \\
      --sheet="CRTA. GUADALMAR" --col-code=C --col-qty=J

  python3 tools/mediciones_validator.py \\
      presupuesto.bc3 mediciones.xlsx \\
      --sheet=Mediciones --col-code=B --col-qty=G --col-unit=C --header=3

Interpretación de los resultados:
  ✓ Cuadran         → código en ambos, cantidad dentro de tolerancia
  ✗ DIFERENCIA      → cantidad distinta (Excel es la fuente de verdad)
  ~ UNIDAD DISTINTA → misma cantidad pero unidad diferente — revisar
  ⚠ Solo en Excel   → partida medida en Excel que no está en bc3
  ⚠ Solo en bc3     → partida en bc3 sin medición en Excel
  - Sin medición    → código en bc3 con cantidad 0 o vacía en ~B
"""

import sys
import os

ENCODING_CANDIDATES = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string
except ImportError:
    print("ERROR: openpyxl no instalado.")
    print("  pip install openpyxl --break-system-packages")
    sys.exit(1)

# Importar bc3_tools desde el mismo directorio
_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _dir)
try:
    from bc3_tools import parse_bc3
except ImportError:
    print("ERROR: bc3_tools.py no encontrado en tools/")
    sys.exit(1)


# ─────────────────────────────────────────────────────────────────────────────
# LECTURA BC3
# ─────────────────────────────────────────────────────────────────────────────

def _parse_bc3_m_quantities(path):
    """
    Lee el archivo bc3 directamente para extraer cantidades de líneas ~M.

    Formatos soportados:
      ~M|PARTIDA|dims|CANTIDAD|desc|           → código simple
      ~M|CAPITULO\\PARTIDA|dims|CANTIDAD|desc| → código con prefijo capítulo

    La cantidad total está en el campo f[2] (índice 2 tras el código).
    Devuelve dict: {codigo_base: qty_float}
    """
    for enc in ENCODING_CANDIDATES:
        try:
            with open(path, encoding=enc) as f:
                lines = f.readlines()
            break
        except (UnicodeDecodeError, LookupError):
            continue
    else:
        return {}

    result = {}
    for line in lines:
        if not line.startswith('~M'):
            continue
        s = line.strip()
        # Quitar ~M| del inicio
        if s.startswith('~M|'):
            s = s[3:]
        elif s.startswith('~M'):
            s = s[2:]
        else:
            continue
        # Quitar | final
        if s.endswith('|'):
            s = s[:-1]
        fields = s.split('|')
        if not fields:
            continue

        raw_code = fields[0]
        # Código base: parte tras la última \ (si contiene prefijo de capítulo)
        base_code = raw_code.split('\\')[-1].strip() if '\\' in raw_code else raw_code.strip()
        if not base_code:
            continue

        # Cantidad total en f[2]
        qty_str = fields[2].strip() if len(fields) > 2 else ''
        try:
            qty = float(qty_str) if qty_str else 0.0
        except ValueError:
            qty = 0.0

        if base_code in result:
            result[base_code] += qty
        else:
            result[base_code] = qty

    return result


def _get_bc3_quantities(bc3, path=None):
    """
    Extrae cantidades del bc3.

    Estrategia (en orden de fiabilidad):
    1. ~B: la fuente más directa cuando existe.
    2. ~M raw: re-parsea el archivo para leer f[2] de cada línea ~M.
       Soporta el formato CAPITULO\\PARTIDA que Presto genera.
    3. ~C sin medición: para detectar partidas que no tienen cantidad en ninguna fuente.

    Devuelve dict: {codigo: {'qty': float, 'unit': str, 'resumen': str, 'fuente': str}}
    Si una partida aparece en varios capítulos (~B o ~M), las cantidades se SUMAN.
    """
    result = {}
    conceptos = bc3['conceptos']

    # 1. Cantidades de ~B
    for cap_code, raw in bc3['rel_caps'].items():
        parts = raw.split('\\')
        for i in range(0, len(parts) - 1, 2):
            code = parts[i].strip()
            if not code:
                continue
            qty_str = parts[i+1].strip() if i+1 < len(parts) else '0'
            try:
                qty = float(qty_str) if qty_str else 0.0
            except ValueError:
                qty = 0.0
            c = conceptos.get(code, {})
            if code not in result:
                result[code] = {
                    'qty': qty,
                    'unit': c.get('unidad', ''),
                    'resumen': c.get('resumen', ''),
                    'fuente': '~B',
                }
            else:
                result[code]['qty'] += qty

    # 2. Cantidades de ~M (re-parseado desde el archivo para obtener f[2])
    m_qtys = _parse_bc3_m_quantities(path) if path else {}
    for base_code, qty in m_qtys.items():
        c = conceptos.get(base_code, {})
        if base_code not in result:
            result[base_code] = {
                'qty': qty,
                'unit': c.get('unidad', ''),
                'resumen': c.get('resumen', ''),
                'fuente': '~M',
            }
        elif result[base_code]['qty'] == 0.0:
            result[base_code]['qty'] = qty
            result[base_code]['fuente'] = '~M'

    # 3. Partidas con solo ~C (sin ~B ni ~M) — para detectar ausencias en Excel
    for code, c in conceptos.items():
        if c['tipo'] not in ('EA', 'EU') and code not in result:
            result[code] = {
                'qty': 0.0,
                'unit': c.get('unidad', ''),
                'resumen': c.get('resumen', ''),
                'fuente': 'sin-medicion',
            }

    return result


# ─────────────────────────────────────────────────────────────────────────────
# LECTURA EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def _col_to_idx(col_str):
    """'B' → 2, 'AA' → 27, etc."""
    try:
        return column_index_from_string(col_str.upper())
    except Exception:
        print(f"ERROR: columna inválida '{col_str}'")
        sys.exit(1)


def _cell_val(cell):
    """Valor de celda como string limpio."""
    v = cell.value
    if v is None:
        return ''
    if isinstance(v, float):
        v = round(v, 10)
        if v == int(v):
            return str(int(v))
        return f"{v:.6f}".rstrip('0').rstrip('.')
    return str(v).strip()


def _get_excel_quantities(path, sheet_name, col_code, col_qty,
                          col_unit=None, col_desc=None, header_row=1):
    """
    Lee las cantidades del Excel.
    Devuelve dict: {codigo: {'qty': float, 'unit': str, 'desc': str, 'fila': int}}
    Filas vacías o con código vacío se ignoran.
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    except Exception as e:
        print(f"ERROR abriendo Excel '{path}': {e}")
        sys.exit(1)

    # Buscar hoja (case-insensitive)
    ws = None
    for name in wb.sheetnames:
        if name.lower() == sheet_name.lower() or name == sheet_name:
            ws = wb[name]
            break
    if ws is None:
        print(f"ERROR: hoja '{sheet_name}' no encontrada en {path}")
        print(f"  Hojas disponibles: {', '.join(wb.sheetnames)}")
        sys.exit(1)

    # Resolver celdas combinadas
    merged_map = {}
    for mr in ws.merged_cells.ranges:
        top_val = _cell_val(ws.cell(row=mr.min_row, column=mr.min_col))
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if not (r == mr.min_row and c == mr.min_col):
                    merged_map[(r, c)] = top_val

    def _get(row_idx, col_idx):
        key = (row_idx, col_idx)
        if key in merged_map:
            return merged_map[key]
        return _cell_val(ws.cell(row=row_idx, column=col_idx))

    cidx_code = _col_to_idx(col_code)
    cidx_qty  = _col_to_idx(col_qty)
    cidx_unit = _col_to_idx(col_unit) if col_unit else None
    cidx_desc = _col_to_idx(col_desc) if col_desc else None

    result = {}
    max_row = ws.max_row or 1

    for row_idx in range(header_row + 1, max_row + 1):
        code = _get(row_idx, cidx_code).strip()
        if not code:
            continue

        qty_str = _get(row_idx, cidx_qty).strip()
        # Normalizar separador decimal español (siempre: punto=miles, coma=decimal)
        qty_str = qty_str.replace('.', '').replace(',', '.')
        try:
            qty = float(qty_str) if qty_str else 0.0
        except ValueError:
            qty = 0.0

        unit = _get(row_idx, cidx_unit).strip() if cidx_unit else ''
        desc = _get(row_idx, cidx_desc).strip() if cidx_desc else ''

        if code in result:
            # Si aparece duplicado, sumar (misma partida en varias filas)
            result[code]['qty'] += qty
        else:
            result[code] = {
                'qty': qty,
                'unit': unit,
                'desc': desc,
                'fila': row_idx,
            }

    return result


# ─────────────────────────────────────────────────────────────────────────────
# CRUCE Y REPORTE
# ─────────────────────────────────────────────────────────────────────────────

def _fmt_qty(q):
    """Formatea cantidad con máximo 4 decimales, sin ceros finales."""
    if q == int(q):
        return f"{int(q):,}"
    return f"{q:,.4f}".rstrip('0').rstrip('.')


def cmd_validate(path_bc3, path_excel, sheet_name,
                 col_code, col_qty, col_unit, col_desc,
                 header_row, tolerance, output_path):

    print(f"\nValidando mediciones:")
    print(f"  bc3   : {path_bc3}")
    print(f"  Excel : {path_excel}  [hoja: {sheet_name}]")
    print(f"  Columnas: código={col_code}  cantidad={col_qty}", end='')
    if col_unit:
        print(f"  unidad={col_unit}", end='')
    if col_desc:
        print(f"  desc={col_desc}", end='')
    print(f"\n  Tolerancia: ±{tolerance}\n")

    bc3 = parse_bc3(path_bc3)
    qtys_bc3   = _get_bc3_quantities(bc3, path_bc3)
    qtys_excel = _get_excel_quantities(
        path_excel, sheet_name, col_code, col_qty,
        col_unit, col_desc, header_row
    )

    codes_bc3   = set(qtys_bc3.keys())
    codes_excel = set(qtys_excel.keys())

    # Excluir partidas tipo 0 sin medición en bc3 que tampoco están en Excel
    # (recursos auxiliares que no tienen medición de obra)
    relevantes_bc3 = {
        c for c in codes_bc3
        if qtys_bc3[c]['fuente'] not in ('sin-medicion',)
        or c in codes_excel
    }

    solo_excel = sorted(codes_excel - codes_bc3)
    solo_bc3   = sorted(relevantes_bc3 - codes_excel)
    comunes    = sorted(relevantes_bc3 & codes_excel)

    cuadran        = []
    diferencias    = []
    unidad_distinta = []
    sin_qty_bc3    = []

    for code in comunes:
        qb = qtys_bc3[code]['qty']
        qe = qtys_excel[code]['qty']
        ub = qtys_bc3[code]['unit'].strip().lower()
        ue = qtys_excel[code]['unit'].strip().lower() if qtys_excel[code]['unit'] else ''

        if qtys_bc3[code]['fuente'] in ('~M-solo', 'sin-medicion'):
            sin_qty_bc3.append(code)
            continue

        diff = abs(qb - qe)
        if diff <= tolerance:
            cuadran.append(code)
            if ue and ub and ue != ub:
                unidad_distinta.append((code, ub, ue))
        else:
            diferencias.append((code, qb, qe, diff,
                                 qtys_bc3[code]['unit'],
                                 qtys_bc3[code]['resumen']))

    # ── Imprimir resumen ──────────────────────────────────────────────────────
    lines = []
    def p(s=''):
        lines.append(s)
        print(s)

    p(f"VALIDACIÓN DE MEDICIONES — {os.path.basename(path_bc3)}")
    p(f"Fecha: {__import__('datetime').date.today().strftime('%d/%m/%Y')}")
    p(f"bc3  : {path_bc3}")
    p(f"Excel: {path_excel}  hoja '{sheet_name}'")
    p()
    p(f"RESUMEN")
    p(f"  Partidas en bc3        : {len(codes_bc3)}")
    p(f"  Partidas en Excel      : {len(codes_excel)}")
    p(f"  Cuadran (±{tolerance})  : {len(cuadran)}")
    p(f"  Con diferencia         : {len(diferencias)}")
    p(f"  Unidad distinta        : {len(unidad_distinta)}")
    p(f"  Solo en Excel          : {len(solo_excel)}")
    p(f"  Solo en bc3            : {len(solo_bc3)}")
    p(f"  Sin medición en bc3~B  : {len(sin_qty_bc3)}")

    if diferencias:
        p()
        p(f"DIFERENCIAS ({len(diferencias)}) — fuente de verdad: Excel")
        p(f"  {'CÓDIGO':<22} {'UD':<5} {'bc3':>12} {'Excel':>12} {'DIFF':>10}  PARTIDA")
        p(f"  {'-'*82}")
        for code, qb, qe, diff, unit, res in sorted(diferencias, key=lambda x: -x[3]):
            p(f"  {code:<22} {unit:<5} {_fmt_qty(qb):>12} {_fmt_qty(qe):>12} "
              f"{_fmt_qty(diff):>10}  {res[:35]}")

    if unidad_distinta:
        p()
        p(f"UNIDAD DISTINTA (misma cantidad) — {len(unidad_distinta)}:")
        for code, ub, ue in unidad_distinta:
            res = qtys_bc3[code]['resumen'][:45]
            p(f"  {code:<22} bc3='{ub}'  Excel='{ue}'  {res}")

    if solo_excel:
        p()
        p(f"SOLO EN EXCEL — no están en bc3 ({len(solo_excel)}):")
        for code in solo_excel[:20]:
            desc = qtys_excel[code].get('desc', '')[:40]
            qty  = _fmt_qty(qtys_excel[code]['qty'])
            p(f"  ⚠ {code:<22} {qty:>10}  {desc}")
        if len(solo_excel) > 20:
            p(f"  ... y {len(solo_excel)-20} más")

    if solo_bc3:
        p()
        p(f"SOLO EN bc3 — sin medición en Excel ({len(solo_bc3)}):")
        for code in solo_bc3[:20]:
            res    = qtys_bc3[code]['resumen'][:40]
            qty    = _fmt_qty(qtys_bc3[code]['qty'])
            fuente = qtys_bc3[code]['fuente']
            p(f"  ⚠ {code:<22} {qty:>10}  [{fuente}]  {res}")
        if len(solo_bc3) > 20:
            p(f"  ... y {len(solo_bc3)-20} más")

    if sin_qty_bc3:
        p()
        p(f"SIN MEDICIÓN EN bc3~B (solo ~M o sin cantidad) — {len(sin_qty_bc3)}:")
        for code in sin_qty_bc3[:10]:
            res = qtys_bc3[code]['resumen'][:45]
            p(f"  - {code:<22} [{qtys_bc3[code]['fuente']}]  {res}")
        if len(sin_qty_bc3) > 10:
            p(f"  ... y {len(sin_qty_bc3)-10} más")
        p(f"  → Verificar ~M de estas partidas manualmente")

    if not diferencias and not solo_excel and not solo_bc3 and not unidad_distinta:
        p()
        p(f"  ✓ Todas las mediciones cuadran")

    if output_path:
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        print(f"\nInforme guardado: {output_path}")

    return diferencias, solo_excel, solo_bc3


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def _parse_args():
    args = sys.argv[1:]
    flags = {}
    positional = []
    for a in args:
        if a.startswith('--'):
            if '=' in a:
                k, v = a[2:].split('=', 1)
                flags[k] = v
            else:
                flags[a[2:]] = True
        else:
            positional.append(a)
    return positional, flags


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(0)

    positional, flags = _parse_args()

    if len(positional) < 2:
        print("Uso: mediciones_validator.py presupuesto.bc3 mediciones.xlsx "
              "--sheet=HOJA --col-code=B --col-qty=H")
        sys.exit(1)

    required = ['sheet', 'col-code', 'col-qty']
    missing = [f"--{r}" for r in required if r not in flags]
    if missing:
        print(f"Faltan parámetros obligatorios: {', '.join(missing)}")
        print("Uso: mediciones_validator.py presupuesto.bc3 mediciones.xlsx "
              "--sheet=HOJA --col-code=B --col-qty=H")
        sys.exit(1)

    cmd_validate(
        path_bc3   = positional[0],
        path_excel = positional[1],
        sheet_name = flags['sheet'],
        col_code   = flags['col-code'],
        col_qty    = flags['col-qty'],
        col_unit   = flags.get('col-unit'),
        col_desc   = flags.get('col-desc'),
        header_row = int(flags.get('header', '1')),
        tolerance  = float(flags.get('tolerance', '0.01')),
        output_path = flags.get('output'),
    )


if __name__ == '__main__':
    main()
