#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
excel_tools.py — Lectura determinista de Excel para obra civil
==============================================================
Lee archivos .xlsx sin LLM: valores exactos, celdas combinadas resueltas,
filas ocultas identificadas, fórmulas devueltas como valor calculado.

REGLA FUNDAMENTAL:
  Este script SOLO LEE. No modifica ningún archivo Excel.
  Los valores numéricos se devuelven tal como los almacena Excel
  (el último valor calculado de la fórmula), sin reinterpretar.

Comandos:
  info    archivo.xlsx
  sheets  archivo.xlsx
  read    archivo.xlsx  HOJA  [--range=A1:Z100]  [--output=salida.csv]
          [--skip-hidden]  [--no-header]
  find    archivo.xlsx  TEXTO  [--sheet=HOJA]

Ejemplos:
  python3 excel_tools.py info 535.2.2\ Control-Calidad.xlsx
  python3 excel_tools.py sheets 535_2_1-GR.xlsx
  python3 excel_tools.py read Control-Calidad.xlsx "Ensayos" --output=ensayos.csv
  python3 excel_tools.py read GR.xlsx "Residuos" --range=A1:H50
  python3 excel_tools.py find Control-Calidad.xlsx "E01.01"
"""

import sys
import os
import csv

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.utils.cell import range_boundaries
except ImportError:
    print("ERROR: openpyxl no instalado.")
    print("  pip install openpyxl --break-system-packages")
    sys.exit(1)


# ─────────────────────────────────────────────────────────────────────────────
# UTILIDADES
# ─────────────────────────────────────────────────────────────────────────────

def _load_wb(path, data_only=True):
    try:
        return openpyxl.load_workbook(path, data_only=data_only, read_only=False)
    except Exception as e:
        print(f"ERROR abriendo '{path}': {e}")
        sys.exit(1)


def _cell_value(cell):
    """
    Devuelve el valor de una celda como string limpio.
    - Números: sin notación científica, hasta 6 decimales significativos
    - Fechas: formato DD/MM/YYYY
    - None / vacío: cadena vacía
    """
    v = cell.value
    if v is None:
        return ''
    if isinstance(v, bool):
        return 'VERDADERO' if v else 'FALSO'
    if isinstance(v, (int, float)):
        # Evitar notación científica y recortar decimales insignificantes
        if isinstance(v, float):
            # Redondear a 10 decimales para eliminar ruido flotante
            v = round(v, 10)
            # Si el resultado es entero, mostrar sin decimales
            if v == int(v):
                return str(int(v))
            # Hasta 6 decimales significativos
            s = f"{v:.6f}".rstrip('0').rstrip('.')
            return s
        return str(v)
    # Fechas
    try:
        import datetime
        if isinstance(v, (datetime.date, datetime.datetime)):
            return v.strftime('%d/%m/%Y')
    except Exception:
        pass
    return str(v).strip()


def _resolve_merged(ws):
    """
    Devuelve un dict {(row, col): valor} con las celdas combinadas expandidas.
    En una celda combinada, solo la celda superior-izquierda tiene valor en
    openpyxl; las demás devuelven None. Propagamos el valor del top-left.
    """
    merged_values = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = (
            merged_range.min_col, merged_range.min_row,
            merged_range.max_col, merged_range.max_row
        )
        top_left = ws.cell(row=min_row, column=min_col)
        val = _cell_value(top_left)
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if not (row == min_row and col == min_col):
                    merged_values[(row, col)] = val
    return merged_values


def _is_row_hidden(ws, row_idx):
    rd = ws.row_dimensions.get(row_idx)
    return rd is not None and rd.hidden


def _parse_range(range_str):
    """'A1:Z100' → (min_col, min_row, max_col, max_row) como enteros"""
    try:
        min_col, min_row, max_col, max_row = range_boundaries(range_str)
        return min_col, min_row, max_col, max_row
    except Exception:
        print(f"ERROR: rango inválido '{range_str}'. Usar formato A1:Z100")
        sys.exit(1)


def _get_sheet(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    # Búsqueda case-insensitive
    for name in wb.sheetnames:
        if name.lower() == sheet_name.lower():
            return wb[name]
    print(f"ERROR: hoja '{sheet_name}' no encontrada.")
    print(f"  Hojas disponibles: {', '.join(wb.sheetnames)}")
    sys.exit(1)


# ─────────────────────────────────────────────────────────────────────────────
# COMANDOS
# ─────────────────────────────────────────────────────────────────────────────

def cmd_info(path):
    wb = _load_wb(path)
    print(f"\n{'='*60}")
    print(f"ARCHIVO: {path}")
    size_kb = os.path.getsize(path) / 1024
    print(f"Tamaño : {size_kb:.1f} KB")
    print(f"Hojas  : {len(wb.sheetnames)}")
    print()

    for name in wb.sheetnames:
        ws = wb[name]
        # Dimensiones reales (ignorar None)
        max_r = ws.max_row or 0
        max_c = ws.max_column or 0

        # Contar filas con datos (no vacías)
        filas_datos = 0
        for row in ws.iter_rows(min_row=1, max_row=max_r):
            if any(c.value is not None for c in row):
                filas_datos += 1

        # Filas ocultas
        ocultas = sum(
            1 for r in range(1, max_r + 1) if _is_row_hidden(ws, r)
        )

        # Celdas combinadas
        n_merged = len(ws.merged_cells.ranges)

        estado = f"{filas_datos} filas con datos"
        if ocultas:
            estado += f", {ocultas} filas ocultas ⚠"
        if n_merged:
            estado += f", {n_merged} rangos combinados"

        print(f"  '{name}'")
        print(f"    Dimensión: {max_r} filas × {max_c} columnas")
        print(f"    {estado}")

    # Rangos con nombre
    try:
        named = list(wb.defined_names)
        if named:
            print(f"\nRangos con nombre: {len(named)}")
            for dn in named[:10]:
                try:
                    print(f"  {dn.name}: {dn.attr_text}")
                except Exception:
                    print(f"  {dn}")
    except Exception:
        pass


def cmd_sheets(path):
    wb = _load_wb(path)
    print(f"\nHojas en '{path}':")
    for i, name in enumerate(wb.sheetnames, 1):
        ws = wb[name]
        print(f"  {i:2}. {name}  "
              f"({ws.max_row}×{ws.max_column})")


def cmd_read(path, sheet_name, range_str=None, output_path=None,
             skip_hidden=False, no_header=False):
    """
    Vuelca una hoja (o rango) a CSV.
    - Celdas combinadas: propaga valor del top-left
    - Filas ocultas: marcadas con *** en la primera columna (o saltadas con --skip-hidden)
    - Fórmulas: devuelve el último valor calculado (data_only=True)
    - Separador: ; (compatible Excel español)
    - Encoding: UTF-8 con BOM
    """
    wb = _load_wb(path, data_only=True)
    ws = _get_sheet(wb, sheet_name)
    merged = _resolve_merged(ws)

    if range_str:
        min_col, min_row, max_col, max_row = _parse_range(range_str)
    else:
        min_row, max_row = 1, ws.max_row or 1
        min_col, max_col = 1, ws.max_column or 1

    if output_path is None:
        base = os.path.splitext(path)[0]
        safe_sheet = sheet_name.replace('/', '-').replace('\\', '-')
        output_path = f"{base}_{safe_sheet}.csv"

    n_rows_written = 0
    n_hidden_skipped = 0
    n_hidden_flagged = 0

    with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f, delimiter=';')

        for row_idx in range(min_row, max_row + 1):
            hidden = _is_row_hidden(ws, row_idx)
            if hidden and skip_hidden:
                n_hidden_skipped += 1
                continue

            row_data = []
            for col_idx in range(min_col, max_col + 1):
                key = (row_idx, col_idx)
                if key in merged:
                    row_data.append(merged[key])
                else:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    row_data.append(_cell_value(cell))

            # Saltar filas completamente vacías (excepto si tienen datos ocultos)
            if not any(row_data):
                continue

            if hidden:
                row_data[0] = f'[OCULTA] {row_data[0]}'
                n_hidden_flagged += 1

            w.writerow(row_data)
            n_rows_written += 1

    rango_info = f" (rango {range_str})" if range_str else ""
    print(f"\nLeído: '{sheet_name}'{rango_info} de {path}")
    print(f"  Filas escritas     : {n_rows_written}")
    if n_hidden_flagged:
        print(f"  Filas ocultas ⚠    : {n_hidden_flagged}  (marcadas con [OCULTA])")
    if n_hidden_skipped:
        print(f"  Filas ocultas      : {n_hidden_skipped}  (saltadas con --skip-hidden)")
    print(f"Exportado: {output_path}")
    print(f"Separador: ';'  |  Encoding: UTF-8 BOM (abre directamente en Excel español)")

    return output_path


def cmd_find(path, texto, sheet_name=None):
    """
    Busca una cadena de texto en todas las celdas del archivo (o en una hoja).
    Case-insensitive. Útil para localizar códigos de partida, descripciones, etc.
    """
    wb = _load_wb(path, data_only=True)
    texto_lower = texto.lower()

    sheets_to_search = [sheet_name] if sheet_name else wb.sheetnames
    resultados = []

    for sname in sheets_to_search:
        ws = _get_sheet(wb, sname)
        for row in ws.iter_rows():
            for cell in row:
                val = _cell_value(cell)
                if texto_lower in val.lower():
                    col_letter = get_column_letter(cell.column)
                    resultados.append({
                        'hoja': sname,
                        'celda': f"{col_letter}{cell.row}",
                        'valor': val[:80],
                        'fila': cell.row,
                        'col': cell.column,
                    })

    if not resultados:
        print(f"\n  '{texto}' no encontrado en {path}")
        return

    print(f"\nResultados para '{texto}' en {path}:")
    print(f"  {'HOJA':<20} {'CELDA':<8} VALOR")
    print(f"  {'-'*65}")
    for r in resultados[:50]:
        print(f"  {r['hoja']:<20} {r['celda']:<8} {r['valor']}")
    if len(resultados) > 50:
        print(f"  ... y {len(resultados)-50} más")

    return resultados


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def _parse_args():
    args = sys.argv[1:]
    flags = {}
    positional = []
    for a in args:
        if a.startswith('--'):
            if '=' in a:
                k, v = a[2:].split('=', 1)
                flags[k] = v
            else:
                flags[a[2:]] = True
        else:
            positional.append(a)
    return positional, flags


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(0)

    positional, flags = _parse_args()
    cmd = positional[0].lower() if positional else ''

    if cmd == 'info':
        if len(positional) < 2:
            print("Uso: excel_tools.py info archivo.xlsx"); sys.exit(1)
        cmd_info(positional[1])

    elif cmd == 'sheets':
        if len(positional) < 2:
            print("Uso: excel_tools.py sheets archivo.xlsx"); sys.exit(1)
        cmd_sheets(positional[1])

    elif cmd == 'read':
        if len(positional) < 3:
            print("Uso: excel_tools.py read archivo.xlsx HOJA [--range=A1:Z100] "
                  "[--output=salida.csv] [--skip-hidden]")
            sys.exit(1)
        cmd_read(
            path        = positional[1],
            sheet_name  = positional[2],
            range_str   = flags.get('range'),
            output_path = flags.get('output'),
            skip_hidden = 'skip-hidden' in flags,
            no_header   = 'no-header' in flags,
        )

    elif cmd == 'find':
        if len(positional) < 3:
            print("Uso: excel_tools.py find archivo.xlsx TEXTO [--sheet=HOJA]")
            sys.exit(1)
        cmd_find(
            path       = positional[1],
            texto      = positional[2],
            sheet_name = flags.get('sheet'),
        )

    else:
        print(f"Comando desconocido: '{cmd}'")
        print("Comandos: info, sheets, read, find")
        sys.exit(1)


if __name__ == '__main__':
    main()
